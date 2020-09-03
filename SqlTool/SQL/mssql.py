import pymssql
import pandas as pd
from threading import Thread,Lock
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import openpyxl
import os

class mssql_client:
    def __init__(self,conninfo):
        self.host=conninfo['HOST']
        self.user=conninfo['USER']
        self.password=conninfo['PASSWORD']
        self.name=conninfo['NAME']
        self.as_dict=conninfo['AS_DICT']        
        self.conn=pymssql.connect(self.host,self.user,self.password,self.name,as_dict=bool(self.as_dict))
        
    def close(self):
        self.conn.close()

    def cursor(self):
        return self.conn.cursor()

    def select(self,tablename):
        cur=self.conn.cursor()
        sql=f"select * from syscolumns where id=object_id('{tablename}')"
        string='select'

        cur.execute(sql)
        columns=[]
        if self.conn.as_dict==1:
            for row in cur.fetchall():
                columns.append(row['name'])
        else:
            for row in cur.fetchall():
                columns.append(row[0])

        string=string+'\t'+'\n\t,'.join(columns)+f'\nfrom {tablename}'        
        cur.close()
        print(string)

    def getDataFrame(self,sql,key:list=None):
        cur=self.conn.cursor()
        cur.execute(sql)
        rs=cur.fetchall()
        des=cur.description
        df=pd.DataFrame(rs,columns=[des[r][0] for r in range(len(des))])

        if key:
            df['rowid']=''
            for col in df.columns:
                if col in key:
                    continue
                df[col]=df[col].astype(str)                               
                df['rowid']=df['rowid']+df[col]            
                
            df=df.set_index(['rowid'])        

        return df

    def getDataFrame2(self,sql,index_col=None):
        cur=self.conn.cursor()
        cur.execute(sql)
        rs=cur.fetchall()
        des=cur.description
        df=pd.DataFrame(rs,columns=[des[r][0] for r in range(len(des))])

        if index_col:            
            df=df.set_index(['index_col'])  
        return df    

    def gap(self,sql1,sql2,key:list=None,prefix=""):
        df1=self.getDataFrame(sql1,key)
        df2=self.getDataFrame(sql2,key)
        for k in key:
            df1[prefix+k]=df2[k]
            df1[f'gap_{k}']=df1[k]-df2[k]                    
                                    
        return df1,df2

    def gap2(self,sql1,sql2,index,key:list=None,prefix=""):
        df1=self.getDataFrame2(sql1,index)
        df2=self.getDataFrame2(sql2,index)
        
        for k in key:
            df1[prefix+k]=df2[k]
            df1[f'gap_{k}']=df1[k]-df2[k]                    
                                    
        return df1,df2    

    def gap_to_excel(self,source:dict,filename):
        '''
        :source dict:
            {sheetname:[df1,df2]}
        
        '''
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        wb=openpyxl.Workbook()
        for sheetname,dfs in source.items():
            df1,df2=dfs
            ws=wb.create_sheet(title=sheetname)
            for r in dataframe_to_rows(df1):
                    ws.append(r)
            min_column=ws.max_column+2
            y=1                    
            for rows in dataframe_to_rows(df2):             
                    for k,v in enumerate(rows,min_column):
                            ws.cell(column=k,row=y,value=v)
                    y+=1

        del wb['Sheet']
        wb.save(filename)
    
class exportConsole():

    def __init__(self,conn,exportRules):
        self.conn=conn
        self.exportRules=exportRules
##        self.gLock=Lock()

    def exportFile(self,data,dirDateTime:str,**kwargs)->str:
        FileName=kwargs.get("FileName")
        ExistsHeader=int(kwargs.get("ExistsHeader"))
        IsZip=int(kwargs.get("IsZip"))
        execType=kwargs.get("execType")        
        DataLocation=kwargs.get("DataLocation")        
        execType=kwargs.get("execType")       
        try:
            wb=openpyxl.load_workbook(DataLocation)
        except FileNotFoundError:
            wb=openpyxl.Workbook()
            
        from openpyxl.utils.dataframe import dataframe_to_rows
        if execType=='ExecuteMany':
            Procs=kwargs.get("Procs")

            for Proc in Procs:
                ws=wb[Proc['SheetName']]
                titles=Proc['titles']
                rn=int(Proc['Rn'])
                
                for i in range(len(titles)):

                    df=data.pop(0)                
                    fill_green=PatternFill(fill_type='solid',fgColor='228B22')

                    if type(df[0])==dict:
                        df=pd.DataFrame(df)
                    
                    if (i+1)%rn==1:
                        ws.append([titles[i]])
                        ws.cell(ws.max_row, 1).fill = fill_green

                        max_row=ws.max_row
                        max_col=len(df.columns)
                        for r in dataframe_to_rows(df,index=False):
                            if r==[None]:
                                continue
                            ws.append(r)
                        ws.append(list())
                    else:
                        dfg=dataframe_to_rows(df,index=False)
                        x = 1
                        ws.cell(row=max_row + x - 1, column=max_col + 2, value=titles[i])
                        ws.cell(row=max_row + x - 1, column=max_col + 2).fill = fill_green

                        while True:
                            try:
                                row_data = dfg.__next__()
                            except StopIteration:
                                break

                            for y in range(1,len(row_data)+1):
                                ws.cell(row=max_row + x, column=max_col + y + 1, value=row_data.pop(0))

                            x+=1
                        max_col += len(df.columns)
        wb.save(os.path.join(dirDateTime,FileName+'.xlsx'))

        print('Done')

        
        if isinstance(data,pd.core.frame.DataFrame)==True:
            pass

    def export(self,ExportDir,*arg):
        parameters=self.exportRules(*arg)
        
        thread_list = []
        print(parameters) 
        for item in parameters:            
            if item['execType']=='ExecuteMany':
                data=[]
                Procs=item["Procs"]
                for Proc in Procs:
                    cur=self.conn.cursor()
                    ProcName=Proc['ProcName']
                    rn=Proc['titles']
                    cur.execute(f"execute {ProcName}")
                    for i in range(len(rn)):
                        data.append(cur.fetchall())
                    cur.close()
        self.exportFile(data,ExportDir,**item)
            # t=Thread(target=self.exportFile,args=(data,ExportDir),kwargs=item)

        for t in thread_list:
            t.start()

        for t in thread_list:
            Thread.join(t)            
        


