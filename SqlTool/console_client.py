import os
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
from threading import Thread,Lock
import xml.etree.ElementTree as ET
from datetime import date,datetime
from shutil import move,copy
import csv

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font,Color
from openpyxl.utils import get_column_letter
import pandas as pd
from sqlalchemy import create_engine, types

from pgtool import SQLClient

year,month=2020,7

class ColumnError(Exception):
    pass

class ExportRules:
    '''控制导出文件的配置    

    '''    
    
    def __init__(self,Rulefile):
        self.tree=ET.parse(Rulefile)
        self.DateFormat_map={'yyyyMMddhhmmss':'%Y%m%d%H%M%S','yyMMddhhmmss':'%y%m%d%H%M%S','yyyyMM':'%Y%m','yyyyMMdd':'%Y%m%d'}
        self.OutputFileName={1:'FileTag',2:'DataLocation',3:'OutputType',4:'OutputDateFormat',5:'DataMonth'}
        self.root=self.tree.getroot()

    def exportRules(self,*arg :str)->list:
        '''
        返回ExportRules.xml文件的配置值
        
        出参说明：
        param Name:
            表、视图或存储过程名称
        param Tag：
            数据查询时间范围
        FileName：
            导出文件名称
        ExistsHeader：
            导出文件是否带表头
        IsZip：
            是否压缩导出文件
        MapptingName：
            匹配名称，同时也是入参值的键值
        Sqlstament:
            可直接运行的SQL语句
        Parameters:
            存储过程参数列表
        '''        
        parameters_list=[]
        
        for MappingName in arg:
            Procs=[]
            
            parameters={}

            try:
                Mapping = self.root.findall(f".//Tables//Table[@MappingName='{MappingName}']")[0]
                FileFormat = Mapping.find('Config//FileFormat').text
                parameters.update(FileFormat=FileFormat,execType='Tables')
            except IndexError:
                try:
                    Mapping=self.root.findall(f".//ExecuteMany//Workbook[@MappingName='{MappingName}']")[0]

                    for childs in Mapping.find('Procs').iter('Proc'):
                        ProcName=childs.get('Name')

                        procs={}
                        temp=[]
                        for child in childs.find('Parameters').iter('Parameter'):
                            Name=child.get('Name')
                            Mode=child.get('Mode')
                            Type=child.get('Type')
                            Value=child.find('.').text
                            temp.append({'Name':Name,'Mode':Mode,'Type':Type,'Value':Value})
                            procs.update(Parameters=temp)

                        titles=[]
                        SheetName=childs.find('Sheet_result_title').get('SheetName')
                        Rn=childs.find('Sheet_result_title').get('Rn')
                        for child in childs.find('Sheet_result_title').iter('title'):
                            titles.append(child.find('.').text)

                        procs.update(ProcName=ProcName,SheetName=SheetName,Rn=Rn,titles=titles)
                        Procs.append(procs)


                    parameters.update(Procs=Procs,execType='ExecuteMany')

                except IndexError:
                    Mapping=self.root.findall(f".//Sqls//Sql[@MappingName='{MappingName}']")[0]

                    for child in Mapping.find('Statements').iter('Statement'):
                        SheetName=child.get('SheetName')
                        Statement=child.find('.').text
                        temp.append({'SheetName':SheetName,'Statement':Statement})
                    parameters.update(Parameters=temp)

            
            Name=Mapping.get('Name')
            Tag=Mapping.get('Tag')
            FileTag=Mapping.find('Config//OutputFileName//FileTag_1').text
            DataLocation=Mapping.find("Config//OutputFileName//DataLocation_2").text
            OutputType=Mapping.find("Config//OutputFileName//OutputType_3").text
            OutputDateFormat=Mapping.find("Config//OutputFileName//OutputDateFormat_4").text
            OutputDateFormat=eval(f"datetime.strftime(datetime.today(),'{self.DateFormat_map[OutputDateFormat]}')") if OutputDateFormat else None
            DataMonth=datetime.strftime(date(year,month,1),'%Y%m') if Mapping.find("Config//OutputFileName//DataMonth_5").text=='1' else None

            CombineFormat=Mapping.find("Config//OutputFileName//CombineFormat").text
            CombineList=[]
            for k in CombineFormat.split('_'):
                    CombineList.append(eval(eval(f"self.OutputFileName[{k}]")))
            FileName='_'.join(CombineList)
            ExistsHeader=Mapping.find("Config//ExistsHeader").text
            
            IsZip=Mapping.find("Config//IsZiped").text
            parameters.update(Name=Name,Tag=Tag,FileName=FileName,ExistsHeader=ExistsHeader,IsZip=IsZip,DataLocation=DataLocation,MappingName=MappingName)
            parameters_list.append(parameters)                

        return parameters_list

class dataConsole(ExportRules):
    def __init__(self,database):
        self.S=SQLClient(database)
        self.conn=self.S.conn
        self.Rulefile=self.S.Rulefile
        if self.Rulefile:
            ExportRules.__init__(self,self.Rulefile)

    def __exportFile(self, datas,dirDateTime:str=None,path: str=None,columns=None, **kwargs) -> str:
        """
        :param datas:
        :param dirDateTime:
        :param path:
        :param columns:
        :param kwargs:
        :return:
        """
        FileName = kwargs.get("FileName")
        ExistsHeader = int(kwargs.get("ExistsHeader"))
        # IsZip = int(kwargs.get("IsZip") if kwargs.get("IsZip") else 0)
        DataLocation = kwargs.get("DataLocation")
        execType = kwargs.get("execType")

        from openpyxl.utils.dataframe import dataframe_to_rows

        if execType == 'Tables':
            data=datas.pop(0)
            if kwargs.get("FileFormat")=='xlsx':
                try:
                    wb = openpyxl.load_workbook(DataLocation)
                except (FileNotFoundError, TypeError):
                    wb = openpyxl.Workbook(write_only=False)

                DEFAULT_FONT = Font(name="Calibri", sz=9, family=2, b=False, i=False,
                                    color=Color(theme=1), scheme="minor")
                wb._fonts.add(DEFAULT_FONT)
                del wb._fonts[0]
                ws = wb.active
                if ExistsHeader:
                    ws.append(columns)
                    ws.freeze_panes = 'A2'
                    for rows in ws.iter_rows(max_row=1):
                        for row in rows:
                            row.font = Font(bold=True)
                maxwidth = [len(columns[k]) for k in range(len(columns))]
                for irow in data:
                    ws.append(irow)
                    for i, l in enumerate(irow):
                        if len(str(l)) > maxwidth[i]:
                            maxwidth[i] = len(str(l))
                for i, w in enumerate(maxwidth):
                    ws.column_dimensions[get_column_letter(i + 1)].width = w + 1
                uploadFile = f'{FileName}.xlsx'
                wb.save(uploadFile)
            else:
                with open(f"{FileName}.csv",'w',newline='') as csvfile:
                    spamwriter = csv.writer(csvfile, dialect='excel')
                    if ExistsHeader:
                        spamwriter.writerow(columns)
                    for row in data:
                        spamwriter.writerow(row)
                uploadFile=f"{FileName}.csv"

        elif execType == 'ExecuteMany':
            try:
                wb = openpyxl.load_workbook(DataLocation)
            except (FileNotFoundError, TypeError):
                wb = openpyxl.Workbook(write_only=False)

            Procs = kwargs.get("Procs")

            for Proc in Procs:
                ws = wb[Proc['SheetName']]
                titles = Proc['titles']
                rn = int(Proc['Rn'])

                for i in range(len(titles)):

                    df = datas.pop(0)
                    fill_green = PatternFill(fill_type='solid', fgColor='228B22')

                    if type(df[0]) == dict:
                        df = pd.DataFrame(df)

                    if (i + 1) % rn == 1:
                        ws.append([titles[i]])
                        ws.cell(ws.max_row, 1).fill = fill_green

                        max_row = ws.max_row
                        max_col = len(df.columns)
                        for r in dataframe_to_rows(df, index=False):
                            if r == [None]:
                                continue
                            ws.append(r)
                        ws.append(list())
                    else:
                        dfg = dataframe_to_rows(df, index=False)
                        x = 1
                        ws.cell(row=max_row + x - 1, column=max_col + 2, value=titles[i])
                        ws.cell(row=max_row + x - 1, column=max_col + 2).fill = fill_green

                        while True:
                            try:
                                row_data = dfg.__next__()
                            except StopIteration:
                                break

                            for y in range(1, len(row_data) + 1):
                                ws.cell(row=max_row + x, column=max_col + y + 1, value=row_data.pop(0))

                            x += 1
                        max_col += len(df.columns)
                        wb.save(os.path.join(FileName + '.xlsx'))


        if path==None :
            if not os.path.exists((ExportPath:=os.path.abspath(f"..\\..\\SQL_Export_Result\\Export_List\\{dirDateTime}"))):
                try:
                    os.makedirs(ExportPath)
                except Exception as err:
                    pass
        else :
            ExportPath=path

        exportFile=os.path.join(ExportPath,uploadFile)
        move(uploadFile,exportFile)

        print('Done')
        return exportFile

    def export(self,*arg,path=None)->str:
        parameters = self.exportRules(*arg)

        thread_list = []
        datas = []
        dirDateTime = datetime.strftime(datetime.now(), '%Y%m%d%H%M%S')
        for item in parameters:
            if item['execType'] == 'Tables':
                with self.conn.cursor() as cur:
                    sql = f"{self.S.select(item['Name'])}"
                    cur.execute(sql)
                    datas.append(cur.fetchall())
                    columns=[k[0] for k in cur.description]

            elif item['execType'] == 'ExecuteMany':

                Procs = item["Procs"]
                for Proc in Procs:
                    cur = self.conn.cursor()
                    ProcName = Proc['ProcName']
                    rn = Proc['titles']
                    cur.execute(f"execute {ProcName}")
                    for i in range(len(rn)):
                        datas.append(cur.fetchall())
                    cur.close()
                columns=None
            t = Thread(target=self.__exportFile, args=(datas,dirDateTime,path,columns), kwargs=item)
            thread_list.append(t)

        for t in thread_list:
            t.start()

        for t in thread_list:
            Thread.join(t)

    def import_from_excel(self,file,dtype=None,**kwargs):
        '''创建表并导入'''

        excel=pd.read_excel(file,dtype=dtype,**kwargs)
        dtypes = {}
        for col in excel.columns:
            if excel[col].dtype == object:
                excel[col] = excel[col].astype(str)
                length=excel[col].apply(lambda x: len(x.encode())).max()
                dtypes[col] = types.VARCHAR(length)
            elif excel[col].dtype == '<M8[ns]': #datetime64[ns]
                dtypes[col] = types.DateTime()
        conn_string = self.S.__str__()
        engine = create_engine(conn_string, encoding='utf-8')
        filename=os.path.splitext(os.path.basename(file))[0]
        with engine.begin():
            excel.to_sql(name=filename,con=engine,if_exists='fail',
                         index=False,chunksize=100,dtype=dtypes)

    def import_from_csv(self,file,tablename,sep=','):
        '''将csv数据导入到已存在的表'''
        columns=self.S.column(tablename)
        try:
            df = pd.read_csv(file,sep=sep,names=columns)
        except UnicodeDecodeError:
            data=[]
            with open(file,newline='', encoding='UTF8', errors="ignore") as csvfile:
                spamreader = csv.reader(csvfile, delimiter=sep)
                for row in spamreader:
                    data.append(row)
            df = pd.DataFrame(data, columns=columns)

        if list(df.columns)!=columns:
            print(list(df.columns))
            raise ColumnError('导入列和表列不对应，无法导入')

        dtypes = self.S.column(tablename, 2)
        for column,dtype in dtypes.items():
            #转换为时间类型，否则oracle写入失败
            if isinstance(dtype,types.DateTime):
                df[column]=pd.to_datetime(df[column])

        conn_string = self.S.__str__()
        engine = create_engine(conn_string, encoding='utf-8')
        with engine.begin():
            df.to_sql(name=tablename, con=engine, if_exists='append',
                         index=False, chunksize=100, dtype=dtypes)

    def sql_to_excel(self,sql,filename,path):
        """查询SQL并导出EXCEL"""
        cur=self.conn.cursor()
        cur.execute(sql)
        columns = [k[0] for k in cur.description]
        datas=[]
        datas.append(cur.fetchall())
        cur.close()
        item={'FileName':filename,'execType':'Tables','FileFormat':'xlsx','ExistsHeader':1}
        return self.__exportFile(datas,path=path,columns=columns,**item)

    @classmethod
    def sync_table(cls,fromdb,todb,table):
        import re
        import pymysql
        ST = SQLClient(todb)
        cur = ST.conn.cursor()
        S = SQLClient(fromdb)
        col1 = S.column(table, 3)
        try:
            #同步表结构
            col2=ST.column(table,3)
            for c in col1:
                if col1[c]!=col2[c]:
                    sql=f'ALTER TABLE {table} MODIFY c {col1[c]};'
                    cur.execute(sql)
        except pymysql.err.ProgrammingError:
            #表不存在，创建表
            sql=S.create(table)
            if sql.find('AUTO_INCREMENT')!=-1:
                sql=re.sub('AUTO_INCREMENT=\d+', 'AUTO_INCREMENT=1', sql)
            cur.execute(sql)
        cur.close()
        S.conn.close()
        ST.conn.close()

    @classmethod
    def sync_data(cls,fromdb,todb,table):
        S=SQLClient(fromdb)
        tconn=SQLClient(todb).conn
        tcur=tconn.cursor()
        for value in S.values(table):
            sql = 'insert into {}\r\nvalues {}'.format(table,value)
            tcur.execute(sql)
            tconn.commit()
        tconn.close()
        print('同步完成')















        


            
        
        
            
            
            
            
        

 
    
